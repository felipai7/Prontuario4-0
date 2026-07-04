#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 Gerador de RELATÓRIO SEMANAL de culturas — export Hermes Pardini (XML em lote)
--------------------------------------------------------------------------------
 Lê o arquivo hpres_*.xml (export de 7 dias do laboratório), extrai TODAS as
 culturas, classifica o status (Positiva / Negativa / Em andamento) e o
 antibiograma, e gera dois relatórios:
     • PDF  (leitura/arquivo)  -> Relatorio_Culturas_<periodo>.pdf
     • XLSX (planilha)         -> Relatorio_Culturas_<periodo>.xlsx

 Uso:
     python gerar_relatorio_semanal.py --xml "hpres_1388_....xml"
     python gerar_relatorio_semanal.py --pasta "D:\\CulturasUTI\\Exports"   (pega o mais recente)

 Sem argumentos, procura o hpres_*.xml mais recente na pasta atual.
 Requisitos: pip install reportlab openpyxl
================================================================================
"""
import os, re, sys, glob, argparse, unicodedata, datetime as dt
import xml.etree.ElementTree as ET

# ------------------------------------------------------------------ util texto
def _norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().upper()

NEG = ["NAO HOUVE CRESCIMENTO", "AUSENCIA DE CRESCIMENTO", "SEM CRESCIMENTO",
       "CULTURA NEGATIVA", "AUSENCIA DE MICRO-ORGANISMOS", "NEGATIVA APOS"]
POS = ["HOUVE CRESCIMENTO", "ISOLADO", "IDENTIFICADO", "POSITIVA"]
PEND = ["EM ANDAMENTO", "AGUARDANDO", "PARCIAL", "A LIBERAR", "SERA LIBERAD"]

# ------------------------------------------------------------------ modelo
class Cultura:
    def __init__(s):
        s.pedido_apoio = s.pedido_lab = s.paciente = ""
        s.material = s.exame = ""
        s.status = "Indefinida"; s.contaminacao = False
        s.resultado_txt = ""
        s.microrganismos = []          # ["Staphylococcus haemolyticus", ...]
        s.antibiograma = []            # [(micro, atb, cim, interpret), ...]
        s.semana = ""                  # rótulo do período (usado no consolidado)

def _txt(el):
    return (el.text or "").strip() if el is not None else ""

# ------------------------------------------------------------------ parser
def parse_xml(path):
    # ElementTree respeita o encoding declarado (ISO-8859-1)
    try:
        root = ET.parse(path).getroot()
    except ET.ParseError:
        # fallback tolerante: lê como latin-1 e limpa caracteres de controle
        raw = open(path, encoding="iso-8859-1", errors="replace").read()
        raw = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", raw)
        root = ET.fromstring(raw.encode("iso-8859-1", "replace"))

    lote = root.find("ControleDeLote")
    periodo = {"ini": "", "fim": "", "emissor": "", "codlab": "", "emissao": ""}
    if lote is not None:
        periodo["emissor"] = _txt(lote.find("Emissor"))
        periodo["codlab"] = _txt(lote.find("CodLab"))
        periodo["emissao"] = _txt(lote.find("DataEmissao"))
        per = lote.find("Periodo")
        if per is not None:
            periodo["ini"] = _txt(per.find("DataInicial"))
            periodo["fim"] = _txt(per.find("DataFinal"))

    culturas = []
    total_pedidos = total_super = 0
    for pedido in root.findall("Pedido"):
        total_pedidos += 1
        nome = _txt(pedido.find("Nome"))
        apoio = _txt(pedido.find("CodPedApoio"))
        lab = _txt(pedido.find("CodPedLab"))
        for sx in pedido.findall("SuperExame"):
            total_super += 1
            if _txt(sx.find("Cultura")) != "1":
                continue
            c = Cultura()
            c.pedido_apoio, c.pedido_lab, c.paciente = apoio, lab, nome
            c.material = _txt(sx.find("MaterialNome"))
            c.exame = _txt(sx.find("ExameNome"))

            valores, isolados = [], []
            for exame in sx.findall("Exame"):
                for item in exame.findall("ItemDeExame"):
                    for res in item.findall("Resultado"):
                        for cont in res.findall("Conteudo"):
                            for v in cont.findall("Valor"):
                                if _txt(v):
                                    valores.append(_txt(v))
                        for iso in res.findall("Isolado"):
                            nome_iso = (iso.get("Nome") or "").strip()
                            if nome_iso:
                                isolados.append(nome_iso)
                            for atb in iso.findall("Antibiotico"):
                                c.antibiograma.append((
                                    nome_iso, (atb.get("Nome") or "").strip(),
                                    (atb.get("CIM") or "").strip(),
                                    (atb.get("Interpretacao") or "").strip()))
            c.resultado_txt = " | ".join(valores)
            c.microrganismos = _dedup(isolados)

            n = _norm(c.resultado_txt)
            if any(k in n for k in NEG) and not isolados:
                c.status = "Negativa"
            elif isolados or ("HOUVE CRESCIMENTO" in n and "NAO HOUVE" not in n):
                c.status = "Positiva"
                if "CONTAMINA" in n:
                    c.contaminacao = True
            elif any(k in n for k in PEND) or not n:
                c.status = "Em andamento"
            else:
                c.status = "Indefinida"
            culturas.append(c)

    return periodo, culturas, total_pedidos, total_super

def _dedup(xs):
    seen, out = set(), []
    for x in xs:
        k = _norm(x)
        if k and k not in seen:
            seen.add(k); out.append(x.strip())
    return out

# ------------------------------------------------------------------ resumo
def resumir(culturas):
    r = {"total": len(culturas), "Positiva": 0, "Negativa": 0,
         "Em andamento": 0, "Indefinida": 0, "contaminacao": 0, "germes": {}}
    for c in culturas:
        r[c.status] = r.get(c.status, 0) + 1
        if c.contaminacao:
            r["contaminacao"] += 1
        for m in c.microrganismos:
            r["germes"][m] = r["germes"].get(m, 0) + 1
    return r

def _abx_resumo(c):
    s = [a for a in c.antibiograma if a[3].lower().startswith("sens")]
    rr = [a for a in c.antibiograma if a[3].lower().startswith("resist")]
    parts = []
    if s:  parts.append("S: " + ", ".join(sorted({a[1] for a in s})))
    if rr: parts.append("R: " + ", ".join(sorted({a[1] for a in rr})))
    return " | ".join(parts)

# ------------------------------------------------------------------ XLSX
def gerar_xlsx(path, periodo, culturas, resumo, incluir_semana=False,
               titulo="Relatório semanal de culturas"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HF = PatternFill("solid", fgColor="1F4E79"); HFONT = Font(bold=True, color="FFFFFF", name="Arial")
    POSF = PatternFill("solid", fgColor="FCE4E4"); NEGF = PatternFill("solid", fgColor="E7F3E7")
    SUB = PatternFill("solid", fgColor="D5E1EF")
    thin = Side(style="thin", color="CCCCCC"); BD = Border(thin, thin, thin, thin)
    def head(ws, cols, ws_widths):
        for i,(h,w) in enumerate(zip(cols, ws_widths),1):
            c = ws.cell(1,i,h); c.fill=HF; c.font=HFONT; c.border=BD
            c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A2"; ws.row_dimensions[1].height=28

    wb = Workbook()
    # ---- Culturas
    ws = wb.active; ws.title="Culturas"
    pre = (["Semana"] if incluir_semana else [])
    cols=pre+["Paciente","Ped. Apoio","Ped. Lab","Material","Exame","Status",
          "Microrganismo(s)","Antibiograma (resumo)","Resultado (texto)"]
    ws_widths=([16] if incluir_semana else [])+[26,12,13,22,26,13,26,34,50]
    head(ws, cols, ws_widths)
    status_col = (7 if incluir_semana else 6)
    ordem={"Positiva":0,"Indefinida":1,"Em andamento":2,"Negativa":3}
    for r,c in enumerate(sorted(culturas,key=lambda x:(x.semana,ordem.get(x.status,9),x.paciente)),2):
        vals=([c.semana] if incluir_semana else [])+[c.paciente,c.pedido_apoio,c.pedido_lab,
              c.material,c.exame,c.status+(" (contaminação)" if c.contaminacao else ""),
              "; ".join(c.microrganismos),_abx_resumo(c),c.resultado_txt]
        for i,v in enumerate(vals,1):
            cell=ws.cell(r,i,v); cell.border=BD; cell.alignment=Alignment(wrap_text=True, vertical="top")
        if c.status=="Positiva": ws.cell(r,status_col).fill=POSF
        elif c.status=="Negativa": ws.cell(r,status_col).fill=NEGF
    # ---- Antibiograma detalhado
    wa=wb.create_sheet("Antibiograma")
    head(wa, pre+["Paciente","Microrganismo","Antibiótico","CIM","Interpretação"],
         ([16] if incluir_semana else [])+[26,28,22,12,16])
    r=2
    for c in culturas:
        for (mic,atb,cim,interp) in c.antibiograma:
            vals=([c.semana] if incluir_semana else [])+[c.paciente,mic,atb,cim,interp]
            for i,v in enumerate(vals,1):
                cell=wa.cell(r,i,v); cell.border=BD
            r+=1
    # ---- Resumo
    wr=wb.create_sheet("Resumo")
    wr.column_dimensions["A"].width=34; wr.column_dimensions["B"].width=18
    wr["A1"]=titulo; wr["A1"].font=Font(bold=True,size=13,color="1F4E79",name="Arial")
    linhas=[("Emissor",periodo["emissor"]),("Cód. laboratório",periodo["codlab"]),
            ("Período (coleta/lote)",f'{periodo["ini"]} a {periodo["fim"]}'),
            ("Data de emissão",periodo["emissao"]),("",""),
            ("Total de culturas",resumo["total"]),("Positivas",resumo["Positiva"]),
            ("  das quais contaminação",resumo["contaminacao"]),
            ("Negativas",resumo["Negativa"]),("Em andamento",resumo["Em andamento"]),
            ("Indefinidas (revisar)",resumo["Indefinida"])]
    for i,(k,v) in enumerate(linhas,start=3):
        a=wr.cell(i,1,k); b=wr.cell(i,2,v)
        if k and not k.startswith(" "): a.font=Font(bold=True,name="Arial")
        if k: a.fill=SUB
    wr.cell(len(linhas)+5,1,"Germes mais frequentes").font=Font(bold=True,color="1F4E79",name="Arial")
    row=len(linhas)+6
    for g,n in sorted(resumo["germes"].items(), key=lambda x:-x[1]):
        wr.cell(row,1,g); wr.cell(row,2,n); row+=1
    wb.save(path)

# ------------------------------------------------------------------ PDF
def gerar_pdf(path, periodo, culturas, resumo, incluir_semana=False,
              titulo="Relatório semanal de culturas microbiológicas"):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, PageBreak)
    ss=getSampleStyleSheet()
    H1=ParagraphStyle("H1",parent=ss["Title"],fontSize=16,textColor=colors.HexColor("#1F4E79"),spaceAfter=4)
    H2=ParagraphStyle("H2",parent=ss["Heading2"],fontSize=12,textColor=colors.HexColor("#2E5496"),spaceBefore=10,spaceAfter=4)
    small=ParagraphStyle("sm",parent=ss["Normal"],fontSize=8,leading=10)
    cell=ParagraphStyle("cell",parent=ss["Normal"],fontSize=7.5,leading=9)
    doc=SimpleDocTemplate(path,pagesize=A4,topMargin=15*mm,bottomMargin=15*mm,
                          leftMargin=12*mm,rightMargin=12*mm,
                          title="Relatório semanal de culturas")
    E=[]
    E.append(Paragraph(titulo,H1))
    E.append(Paragraph(f'{periodo["emissor"]} — Cód. lab {periodo["codlab"]} · '
                       f'Período {periodo["ini"]} a {periodo["fim"]} · '
                       f'Emissão {periodo["emissao"]}',small))
    E.append(Spacer(1,6))
    # Painel de resumo
    resumo_tab=[["Total culturas","Positivas","Negativas","Em andamento","Indefinidas"],
                [resumo["total"],resumo["Positiva"],resumo["Negativa"],
                 resumo["Em andamento"],resumo["Indefinida"]]]
    t=Table(resumo_tab,colWidths=[37*mm]*5)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTSIZE",(0,0),(-1,-1),9),("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#CCCCCC")),
        ("BACKGROUND",(1,1),(1,1),colors.HexColor("#FCE4E4")),
        ("BACKGROUND",(2,1),(2,1),colors.HexColor("#E7F3E7")),
        ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
    E.append(t)
    if resumo["contaminacao"]:
        E.append(Paragraph(f'Obs.: {resumo["contaminacao"]} positiva(s) sinalizada(s) como provável contaminação.',small))

    # Germes frequentes
    if resumo["germes"]:
        E.append(Paragraph("Microrganismos isolados na semana",H2))
        g=[["Microrganismo","Nº"]]+[[k,v] for k,v in sorted(resumo["germes"].items(),key=lambda x:-x[1])]
        tg=Table(g,colWidths=[130*mm,20*mm])
        tg.setStyle(_ts_head())
        E.append(tg)

    # Positivas com antibiograma
    pos=[c for c in culturas if c.status=="Positiva"]
    E.append(Paragraph(f"Culturas POSITIVAS ({len(pos)})",H2))
    if pos:
        hdr=(["Semana"] if incluir_semana else [])+["Paciente","Material / Exame",
             "Microrganismo","Antibiograma (S / R)"]
        data=[hdr]
        for c in pos:
            linha=([Paragraph(c.semana,cell)] if incluir_semana else [])+[
                         Paragraph(c.paciente,cell),
                         Paragraph(f"{c.material}<br/><font size=6 color='#666666'>{c.exame}</font>",cell),
                         Paragraph("; ".join(c.microrganismos) or ("(contaminação)" if c.contaminacao else "—"),cell),
                         Paragraph(_abx_resumo(c) or "—",cell)]
            data.append(linha)
        cw=([22*mm] if incluir_semana else [])+([32*mm,38*mm,36*mm,58*mm] if incluir_semana else [38*mm,45*mm,40*mm,63*mm])
        tp=Table(data,colWidths=cw,repeatRows=1)
        tp.setStyle(_ts_head(rosa=True))
        E.append(tp)
    else:
        E.append(Paragraph("Nenhuma cultura positiva no período.",small))

    # Em andamento / indefinidas
    pend=[c for c in culturas if c.status in ("Em andamento","Indefinida")]
    if pend:
        E.append(Paragraph(f"Em andamento / a revisar ({len(pend)})",H2))
        data=[["Paciente","Material / Exame","Status","Resultado"]]
        for c in pend:
            data.append([Paragraph(c.paciente,cell),Paragraph(f"{c.material} — {c.exame}",cell),
                         Paragraph(c.status,cell),Paragraph(c.resultado_txt or "—",cell)])
        tpe=Table(data,colWidths=[38*mm,55*mm,25*mm,68*mm],repeatRows=1)
        tpe.setStyle(_ts_head()); E.append(tpe)

    # Negativas (lista compacta)
    neg=[c for c in culturas if c.status=="Negativa"]
    E.append(Paragraph(f"Culturas negativas ({len(neg)})",H2))
    if neg:
        data=[["Paciente","Material / Exame"]]
        for c in neg:
            data.append([Paragraph(c.paciente,cell),Paragraph(f"{c.material} — {c.exame}",cell)])
        tn=Table(data,colWidths=[60*mm,126*mm],repeatRows=1)
        tn.setStyle(_ts_head(verde=True)); E.append(tn)

    E.append(Spacer(1,8))
    E.append(Paragraph("Documento gerado automaticamente a partir do export em lote do laboratório. "
                       "Contém dados sensíveis de saúde — uso restrito (LGPD).",
                       ParagraphStyle("f",parent=small,textColor=colors.HexColor("#888888"))))
    doc.build(E, onFirstPage=_rodape, onLaterPages=_rodape)

def _ts_head(rosa=False,verde=False):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    head=colors.HexColor("#1F4E79")
    st=[("BACKGROUND",(0,0),(-1,0),head),("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTSIZE",(0,0),(-1,0),8),("FONTSIZE",(0,1),(-1,-1),7.5),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#CCCCCC")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F7FA")])]
    if rosa: st.append(("LINEBELOW",(0,0),(-1,0),1,colors.HexColor("#C00000")))
    if verde: st.append(("LINEBELOW",(0,0),(-1,0),1,colors.HexColor("#2E7D32")))
    return TableStyle(st)

def _rodape(canvas, doc):
    from reportlab.lib.units import mm
    canvas.saveState(); canvas.setFont("Helvetica",7)
    canvas.setFillColorRGB(.53,.53,.53)
    canvas.drawRightString(200*mm,8*mm,f"Página {doc.page}")
    canvas.drawString(12*mm,8*mm,"Confidencial — dados de saúde")
    canvas.restoreState()

# ------------------------------------------------------------------ main
def listar_xmls(pasta):
    cs=glob.glob(os.path.join(pasta,"hpres_*.xml")) or glob.glob(os.path.join(pasta,"*.xml"))
    return sorted(cs, key=os.path.getmtime)

def periodo_tag(periodo):
    return f'{periodo["ini"]}_a_{periodo["fim"]}' if periodo["ini"] else dt.date.today().isoformat()

def processar_arquivo(xml, saida, pular_existentes=False):
    """Gera o relatório de UM arquivo. Retorna (periodo, culturas, resumo, tag)."""
    periodo, culturas, tp, ts = parse_xml(xml)
    resumo = resumir(culturas)
    tag = periodo_tag(periodo)
    for c in culturas:
        c.semana = tag.replace("_a_", " a ")
    pdf = os.path.join(saida, f"Relatorio_Culturas_{tag}.pdf")
    xlsx = os.path.join(saida, f"Relatorio_Culturas_{tag}.xlsx")
    if pular_existentes and os.path.exists(pdf) and os.path.exists(xlsx):
        print(f"[pulado] {tag} (já existe)")
        return periodo, culturas, resumo, tag
    gerar_pdf(pdf, periodo, culturas, resumo)
    gerar_xlsx(xlsx, periodo, culturas, resumo)
    print(f"[ok] {os.path.basename(xml)} -> {tag} | culturas {resumo['total']} "
          f"| positivas {resumo['Positiva']} (contam {resumo['contaminacao']}) "
          f"| negativas {resumo['Negativa']} | andamento {resumo['Em andamento']} "
          f"| indefinidas {resumo['Indefinida']}")
    return periodo, culturas, resumo, tag

def consolidar(entradas, saida):
    """entradas: lista de (periodo, culturas). Gera um relatório único do intervalo total."""
    todas=[]; inis=[]; fims=[]; emissor=codlab=""
    for periodo, culturas in entradas:
        todas.extend(culturas)
        if periodo.get("ini"): inis.append(periodo["ini"])
        if periodo.get("fim"): fims.append(periodo["fim"])
        emissor=emissor or periodo.get("emissor",""); codlab=codlab or periodo.get("codlab","")
    if not todas:
        print("Nada a consolidar."); return
    per={"ini":min(inis) if inis else "", "fim":max(fims) if fims else "",
         "emissor":emissor, "codlab":codlab, "emissao":f"consolidado de {len(entradas)} semana(s)"}
    resumo=resumir(todas)
    tag=f'{per["ini"]}_a_{per["fim"]}' if per["ini"] else dt.date.today().isoformat()
    titulo="Consolidado retrospectivo de culturas"
    pdf=os.path.join(saida, f"CONSOLIDADO_Culturas_{tag}.pdf")
    xlsx=os.path.join(saida, f"CONSOLIDADO_Culturas_{tag}.xlsx")
    gerar_pdf(pdf, per, todas, resumo, incluir_semana=True, titulo=titulo)
    gerar_xlsx(xlsx, per, todas, resumo, incluir_semana=True, titulo=titulo)
    print(f"[CONSOLIDADO] {tag} | semanas {len(entradas)} | culturas {resumo['total']} "
          f"| positivas {resumo['Positiva']} | negativas {resumo['Negativa']}")
    print("PDF :", pdf); print("XLSX:", xlsx)

def main():
    ap=argparse.ArgumentParser(description="Gerador de relatório de culturas (semanal ou retroativo).")
    ap.add_argument("--xml", help="arquivo XML específico")
    ap.add_argument("--pasta", default=".", help="pasta com os exports (para --lote ou XML mais recente)")
    ap.add_argument("--saida", help="pasta de saída (default: <pasta>/Relatorios)")
    ap.add_argument("--lote", action="store_true",
                    help="RETROATIVO: processa TODOS os hpres_*.xml da pasta (um relatório por semana)")
    ap.add_argument("--consolidar", action="store_true",
                    help="também gera um relatório único consolidado do período inteiro")
    ap.add_argument("--pular-existentes", action="store_true",
                    help="no modo lote, não regenera semanas que já têm relatório")
    a=ap.parse_args()

    if a.lote:
        arquivos=listar_xmls(a.pasta)
        base=os.path.abspath(a.pasta)
    elif a.xml:
        arquivos=[a.xml]; base=os.path.dirname(os.path.abspath(a.xml))
    else:
        arquivos=listar_xmls(a.pasta)[-1:]; base=os.path.abspath(a.pasta)
    if not arquivos:
        print("Nenhum XML encontrado."); sys.exit(1)

    saida=a.saida or os.path.join(base,"Relatorios")
    os.makedirs(saida, exist_ok=True)

    entradas=[]
    for xml in arquivos:
        if not os.path.exists(xml):
            print(f"[ignorado] não existe: {xml}"); continue
        try:
            periodo, culturas, resumo, tag = processar_arquivo(xml, saida, a.pular_existentes)
            entradas.append((periodo, culturas))
        except Exception as e:
            print(f"[ERRO] {os.path.basename(xml)}: {type(e).__name__}: {e}")

    print(f"\nTotal de semanas processadas: {len(entradas)}")
    if a.consolidar:
        consolidar(entradas, saida)

if __name__ == "__main__":
    main()
