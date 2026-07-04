#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
 Automação de extração de CULTURAS POSITIVAS de pacientes de UTI (18 meses)
 Portais: My Pardini e DB (acesso web, sem API)
--------------------------------------------------------------------------------
 Este é um MOTOR DE REFERÊNCIA, comentado e resiliente. Ele NÃO roda "de fábrica":
 os seletores reais das telas só existem quando você inspeciona os portais com
 acesso autorizado. Procure os marcadores  # >>> CONFIGURAR  e preencha.

 Pode ser usado de duas formas:
   1) Autônomo:   python automacao_culturas_selenium.py --planilha pacientes.xlsx
   2) Chamado pelo Power Automate Desktop (PAD) para a parte web/parse.

 Instalação (uma vez):
   pip install selenium webdriver-manager pdfplumber openpyxl
   (para OCR de contingência, opcional: pip install pytesseract pillow  + Tesseract)

 Conformidade: confirme os termos de uso do portal e a LGPD antes de rodar.
================================================================================
"""

import os, sys, csv, json, time, hashlib, logging, argparse, datetime as dt
from dataclasses import dataclass, field

# --- dependências externas (import tardio para dar mensagem amigável) ---------
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, WebDriverException
except ImportError:
    print("Instale as dependências: pip install selenium webdriver-manager pdfplumber openpyxl")
    raise

# ==============================================================================
# 1. CONFIGURAÇÃO (ajuste aqui — tudo centralizado para manutenção fácil)
# ==============================================================================
CONFIG = {
    "raiz":            r"D:\CulturasUTI",           # >>> CONFIGURAR pasta raiz
    "janela_meses":    18,
    "timeout":         30,                           # espera máxima por elemento (s)
    "tentativas_max":  3,
    "backoff":         [5, 15, 30],                  # espera entre tentativas (s)
    "sistema":         "my_pardini",                 # "my_pardini" ou "db"
    "url_login":       "https://resultados.grupofleury.com.br/",  # >>> CONFIRMAR
    "termos_cultura":  ["cultura", "hemocultura", "urocultura", "coprocultura",
                         "cultura de secre", "ponta de cateter", "cultura de vigil"],
    "termos_positivo": ["positivo", "cresc", "isolado", "identificado"],  # em texto livre
    "termos_negativo": ["negativo", "ausência de cresc", "sem cresc", "nao houve cresc"],
}

# Seletores por sistema — >>> CONFIGURAR com o que você inspecionar nas telas.
SELETORES = {
    "my_pardini": {
        "campo_usuario": (By.ID, "CONFIGURAR_usuario"),
        "campo_senha":   (By.ID, "CONFIGURAR_senha"),
        "btn_entrar":    (By.XPATH, "//button[contains(.,'Entrar')]"),
        "campo_busca":   (By.ID, "CONFIGURAR_busca"),
        "btn_buscar":    (By.XPATH, "//button[contains(.,'Buscar')]"),
        "lista_exames":  (By.CSS_SELECTOR, "CONFIGURAR_lista"),
        "link_laudo":    (By.XPATH, "CONFIGURAR_link_pdf"),
        "flag_login":    (By.ID, "CONFIGURAR_usuario"),  # se visível => sessão caiu
    },
    "db": {  # >>> CONFIGURAR análogo ao acima para o sistema DB
        "campo_usuario": (By.ID, "CONFIGURAR"),
        "campo_senha":   (By.ID, "CONFIGURAR"),
        "btn_entrar":    (By.XPATH, "CONFIGURAR"),
        "campo_busca":   (By.ID, "CONFIGURAR"),
        "btn_buscar":    (By.XPATH, "CONFIGURAR"),
        "lista_exames":  (By.CSS_SELECTOR, "CONFIGURAR"),
        "link_laudo":    (By.XPATH, "CONFIGURAR"),
        "flag_login":    (By.ID, "CONFIGURAR"),
    },
}

# ==============================================================================
# 2. INFRAESTRUTURA: pastas, logs, checkpoint
# ==============================================================================
def caminhos():
    raiz = CONFIG["raiz"]
    p = {
        "controle":   os.path.join(raiz, "_controle"),
        "consolidado":os.path.join(raiz, "_consolidado"),
        "pacientes":  os.path.join(raiz, "Pacientes"),
    }
    p["screenshots"] = os.path.join(p["controle"], "screenshots")
    for d in p.values():
        os.makedirs(d, exist_ok=True)
    p["checkpoint"]  = os.path.join(p["controle"], "checkpoint.json")
    p["log"]         = os.path.join(p["controle"], "log_execucao.csv")
    p["erros"]       = os.path.join(p["controle"], "erros.csv")
    p["incompletos"] = os.path.join(p["controle"], "pacientes_incompletos.csv")
    p["consol_csv"]  = os.path.join(p["consolidado"], "consolidado.csv")
    return p

def setup_log(path_log):
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.FileHandler(path_log.replace(".csv", ".txt"), encoding="utf-8"),
                  logging.StreamHandler(sys.stdout)],
    )

def append_csv(path, row, header=None):
    novo = not os.path.exists(path)
    with open(path, "a", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        if novo and header:
            w.writerow(header)
        w.writerow(row)

def carregar_checkpoint(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {"concluidos": [], "contadores": {"processados": 0, "positivos": 0,
                                             "pdfs": 0, "erros": 0}}

def salvar_checkpoint(path, ck):
    ck["atualizado_em"] = dt.datetime.now().isoformat(timespec="seconds")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(ck, f, ensure_ascii=False, indent=2)

# ==============================================================================
# 3. UTILITÁRIOS DE ROBUSTEZ (espera, retry, validação, hash, nome de arquivo)
# ==============================================================================
def esperar(driver, seletor, timeout=None):
    """Espera EXPLÍCITA por elemento — nunca sleep fixo."""
    timeout = timeout or CONFIG["timeout"]
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located(seletor))

def esperar_clicavel(driver, seletor, timeout=None):
    timeout = timeout or CONFIG["timeout"]
    return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(seletor))

def com_retry(func, descricao, tentativas=None, backoff=None):
    """Executa uma ação de rede com retentativas e backoff exponencial."""
    tentativas = tentativas or CONFIG["tentativas_max"]
    backoff = backoff or CONFIG["backoff"]
    ultima = None
    for i in range(tentativas):
        try:
            return func()
        except (TimeoutException, WebDriverException) as e:
            ultima = e
            espera = backoff[min(i, len(backoff) - 1)]
            logging.warning("Falha em '%s' (tentativa %d/%d): %s. Aguardando %ds.",
                            descricao, i + 1, tentativas, type(e).__name__, espera)
            time.sleep(espera)
    raise ultima

def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for bloco in iter(lambda: f.read(8192), b""):
            h.update(bloco)
    return h.hexdigest()

def validar_pdf(path):
    """Retorna (ok, motivo)."""
    if not os.path.exists(path):
        return False, "arquivo inexistente"
    if os.path.getsize(path) < 10 * 1024:
        return False, "tamanho suspeito (<10KB) — possível corrompido"
    with open(path, "rb") as f:
        if f.read(5) != b"%PDF-":
            return False, "cabeçalho não é PDF"
    return True, "ok"

def nome_arquivo(paciente_id, nome, data_coleta, tipo, atendimento, seq):
    def limpa(s):
        return "".join(c for c in str(s) if c.isalnum() or c in "-_").upper()[:20]
    return f"{limpa(paciente_id)}_{limpa(nome)}_{data_coleta}_{limpa(tipo)}_{limpa(atendimento)}_{seq:02d}.pdf"

def classificar_resultado(texto):
    """Heurística: positivo se houver termo de crescimento e não for claramente negativo."""
    t = (texto or "").lower()
    if any(n in t for n in CONFIG["termos_negativo"]):
        return "negativo"
    if any(p in t for p in CONFIG["termos_positivo"]):
        return "positivo"
    return "indefinido"   # marcar para revisão humana

# ==============================================================================
# 4. NAVEGADOR / SESSÃO
# ==============================================================================
def abrir_navegador(pasta_download):
    opts = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": pasta_download,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,  # baixa o PDF em vez de abrir no visualizador
    }
    opts.add_experimental_option("prefs", prefs)
    # opts.add_argument("--headless=new")  # deixe VISÍVEL para tratar 2FA no login
    driver = webdriver.Chrome(options=opts)
    driver.set_page_load_timeout(CONFIG["timeout"] * 2)
    return driver

def login(driver):
    sel = SELETORES[CONFIG["sistema"]]
    driver.get(CONFIG["url_login"])
    # >>> CONFIGURAR: se houver 2FA, pause aqui e peça intervenção humana:
    #   input("Faça o login/2FA manualmente e tecle ENTER para continuar...")
    try:
        esperar(driver, sel["campo_usuario"])
        driver.find_element(*sel["campo_usuario"]).send_keys(os.environ.get("LAB_USER", ""))
        driver.find_element(*sel["campo_senha"]).send_keys(os.environ.get("LAB_PASS", ""))
        esperar_clicavel(driver, sel["btn_entrar"]).click()
    except TimeoutException:
        logging.info("Tela de login não detectada automaticamente — talvez já autenticado.")
    logging.info("Login concluído (ou já autenticado).")

def sessao_expirada(driver):
    sel = SELETORES[CONFIG["sistema"]]
    try:
        driver.find_element(*sel["flag_login"])
        return True      # elemento de login visível => caiu a sessão
    except WebDriverException:
        return False

# ==============================================================================
# 5. PROCESSAMENTO DE UM PACIENTE
# ==============================================================================
@dataclass
class Resultado:
    paciente_id: str
    nome: str = ""
    atendimento: str = ""
    status: str = "OK"
    positivos: int = 0
    pdfs: int = 0
    linhas: list = field(default_factory=list)  # dicts p/ o consolidado
    obs: str = ""

def data_dentro_janela(data_coleta, data_limite):
    try:
        d = dt.datetime.strptime(data_coleta, "%Y-%m-%d").date()
        return d >= data_limite
    except Exception:
        return True  # em dúvida, não descartar

def processar_paciente(driver, paciente, data_limite, paths, hashes_conhecidos):
    """
    paciente: dict com chaves como id, nome, atendimento.
    Esta função contém a LÓGICA; os pontos >>> CONFIGURAR dependem da tela real.
    """
    sel = SELETORES[CONFIG["sistema"]]
    res = Resultado(paciente_id=paciente.get("id", ""),
                    nome=paciente.get("nome", ""),
                    atendimento=paciente.get("atendimento", ""))

    # 5.1 pesquisar (por atendimento se houver; senão por id/nome)
    chave = paciente.get("atendimento") or paciente.get("id") or paciente.get("nome")
    def _buscar():
        campo = esperar(driver, sel["campo_busca"])
        campo.clear(); campo.send_keys(str(chave))
        esperar_clicavel(driver, sel["btn_buscar"]).click()
        esperar(driver, sel["lista_exames"])
    try:
        com_retry(_buscar, f"buscar {chave}")
    except TimeoutException:
        res.status = "não encontrado"
        append_csv(paths["erros"], [dt.datetime.now().isoformat(), res.paciente_id,
                   "busca", "não encontrado", str(chave)],
                   header=["timestamp","paciente","etapa","tipo","detalhe"])
        return res

    # 5.2 percorrer exames (>>> CONFIGURAR extração real da lista)
    # Estrutura ilustrativa: cada 'linha_exame' teria data, nome_exame, url_laudo.
    linhas_exame = extrair_lista_exames(driver)   # ver função abaixo (placeholder)

    seq = 0
    for ex in linhas_exame:
        eh_cultura = any(t in ex["nome"].lower() for t in CONFIG["termos_cultura"])
        if not eh_cultura:
            continue
        if not data_dentro_janela(ex["data"], data_limite):
            continue

        # 5.3 abrir/inspecionar laudo e classificar
        resultado = classificar_resultado(ex.get("texto_resumo", ""))
        if resultado == "negativo":
            continue                      # ignora negativos
        seq += 1
        res.positivos += 1

        # 5.4 baixar PDF
        destino_dir = os.path.join(paths["pacientes"],
                                   f"{res.paciente_id}_{res.nome}".strip("_"),
                                   res.atendimento or "sem_atendimento")
        os.makedirs(destino_dir, exist_ok=True)
        fname = nome_arquivo(res.paciente_id, res.nome, ex["data"],
                             ex["nome"], res.atendimento, seq)
        destino = os.path.join(destino_dir, fname)

        if os.path.exists(destino):
            logging.info("PDF já existe, pulando download: %s", fname)
        else:
            try:
                com_retry(lambda: baixar_laudo(driver, ex, destino_dir, destino),
                          f"download {fname}")
            except Exception as e:
                append_csv(paths["erros"], [dt.datetime.now().isoformat(), res.paciente_id,
                           "download", type(e).__name__, fname],
                           header=["timestamp","paciente","etapa","tipo","detalhe"])
                res.obs += f"falha download {fname}; "
                continue

        # 5.5 validar
        ok, motivo = validar_pdf(destino)
        if not ok:
            append_csv(paths["erros"], [dt.datetime.now().isoformat(), res.paciente_id,
                       "validação", motivo, fname],
                       header=["timestamp","paciente","etapa","tipo","detalhe"])
            res.obs += f"pdf inválido ({motivo}); "
            continue

        # 5.6 deduplicação por hash
        h = sha256(destino)
        if h in hashes_conhecidos:
            logging.info("Duplicado por hash, removendo cópia: %s", fname)
            os.remove(destino)
            continue
        hashes_conhecidos.add(h)
        res.pdfs += 1

        # 5.7 parse do PDF → campos do consolidado
        campos = parse_laudo_pdf(destino)
        linha = {
            "nome": res.nome, "identificador": res.paciente_id,
            "atendimento": res.atendimento, "data_coleta": ex["data"],
            "tipo_cultura": ex["nome"], "material": campos.get("material", ""),
            "microrganismo": campos.get("microrganismo", ""),
            "resultado": "Positivo", "antibiograma": campos.get("antibiograma", ""),
            "caminho_pdf": os.path.relpath(destino, CONFIG["raiz"]),
            "data_download": dt.date.today().isoformat(),
            "status": "OK", "observacoes": campos.get("obs", ""),
        }
        res.linhas.append(linha)

    return res

# ------ Placeholders a preencher com a tela real ------------------------------
def extrair_lista_exames(driver):
    """
    >>> CONFIGURAR: leia as linhas da tabela de exames e devolva uma lista de dicts:
        [{"data":"2026-03-14","nome":"Hemocultura","texto_resumo":"...","url_laudo":"..."}]
    Use driver.find_elements(*SELETORES[...]["lista_exames"]) e extraia os campos.
    """
    return []

def baixar_laudo(driver, exame, destino_dir, destino_final):
    """
    >>> CONFIGURAR: clicar no link/botão do laudo, aguardar o download concluir
    (sem arquivos .crdownload) e renomear para 'destino_final'.
    """
    antes = set(os.listdir(destino_dir))
    esperar_clicavel(driver, SELETORES[CONFIG["sistema"]]["link_laudo"]).click()
    # aguardar conclusão do download
    fim = time.time() + CONFIG["timeout"]
    while time.time() < fim:
        novos = set(os.listdir(destino_dir)) - antes
        prontos = [n for n in novos if not n.endswith(".crdownload") and not n.endswith(".part")]
        if prontos:
            os.rename(os.path.join(destino_dir, prontos[0]), destino_final)
            return
        time.sleep(1)
    raise TimeoutException("download não concluído")

def parse_laudo_pdf(path):
    """Extrai campos do laudo. Usa pdfplumber; OCR só se não houver texto."""
    campos = {"material": "", "microrganismo": "", "antibiograma": "", "obs": ""}
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            texto = "\n".join((pg.extract_text() or "") for pg in pdf.pages)
        if not texto.strip():
            campos["obs"] = "sem camada de texto — considerar OCR"
            return campos
        # >>> CONFIGURAR: regras/regex conforme o layout do laudo do laboratório.
        campos["_texto"] = texto[:2000]
    except Exception as e:
        campos["obs"] = f"parse falhou: {type(e).__name__}"
    return campos

# ==============================================================================
# 6. ORQUESTRAÇÃO (quando rodar autônomo)
# ==============================================================================
def ler_pacientes(xlsx_path):
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    cabec = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    pacientes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {cabec[i]: (row[i] if i < len(row) else None) for i in range(len(cabec))}
        pacientes.append({
            "id": d.get("identificador") or d.get("id") or "",
            "nome": d.get("nome") or d.get("paciente") or "",
            "atendimento": d.get("atendimento") or "",
        })
    return pacientes

def campos_obrigatorios_ok(p):
    return bool((p.get("id") or p.get("atendimento")) and p.get("nome"))

def gravar_consolidado(paths, linhas):
    header = ["nome","identificador","atendimento","data_coleta","tipo_cultura","material",
              "microrganismo","resultado","antibiograma","caminho_pdf","data_download",
              "status","observacoes"]
    for l in linhas:
        append_csv(paths["consol_csv"], [l.get(k, "") for k in header], header=header)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", required=True, help="Excel com os pacientes")
    ap.add_argument("--sistema", default=CONFIG["sistema"], choices=["my_pardini", "db"])
    args = ap.parse_args()
    CONFIG["sistema"] = args.sistema

    paths = caminhos()
    setup_log(paths["log"])
    inicio = time.time()
    ck = carregar_checkpoint(paths["checkpoint"])
    concluidos = set(ck["concluidos"])
    data_limite = (dt.date.today() - dt.timedelta(days=int(CONFIG["janela_meses"] * 30.44)))
    hashes = set()

    pacientes = ler_pacientes(args.planilha)
    logging.info("Pacientes lidos: %d | janela desde %s", len(pacientes), data_limite)

    driver = abrir_navegador(paths["pacientes"])
    try:
        login(driver)
        for p in pacientes:
            chave = p.get("id") or p.get("atendimento")
            if chave in concluidos:
                continue
            if not campos_obrigatorios_ok(p):
                append_csv(paths["incompletos"], [p.get("id"), p.get("nome"), p.get("atendimento")],
                           header=["id","nome","atendimento"])
                continue
            # laço tolerante a falhas com recuperação de sessão
            for tentativa in range(1, CONFIG["tentativas_max"] + 1):
                try:
                    res = processar_paciente(driver, p, data_limite, paths, hashes)
                    gravar_consolidado(paths, res.linhas)
                    ck["contadores"]["processados"] += 1
                    ck["contadores"]["positivos"] += res.positivos
                    ck["contadores"]["pdfs"] += res.pdfs
                    if res.status not in ("OK", "não encontrado"):
                        ck["contadores"]["erros"] += 1
                    concluidos.add(chave)
                    ck["concluidos"] = list(concluidos)
                    salvar_checkpoint(paths["checkpoint"], ck)
                    break
                except Exception as e:
                    driver.save_screenshot(os.path.join(paths["screenshots"],
                                                        f"{chave}_{tentativa}.png"))
                    append_csv(paths["erros"], [dt.datetime.now().isoformat(), chave,
                               "processar", type(e).__name__, str(e)[:200]],
                               header=["timestamp","paciente","etapa","tipo","detalhe"])
                    if sessao_expirada(driver):
                        logging.warning("Sessão expirada — refazendo login.")
                        login(driver)
                        continue
                    if tentativa == CONFIG["tentativas_max"]:
                        ck["contadores"]["erros"] += 1
                        logging.error("Paciente %s falhou após %d tentativas — seguindo.",
                                      chave, tentativa)
    finally:
        driver.quit()

    dur = time.time() - inicio
    relatorio = {
        "processados": ck["contadores"]["processados"],
        "culturas_positivas": ck["contadores"]["positivos"],
        "pdfs_baixados": ck["contadores"]["pdfs"],
        "erros": ck["contadores"]["erros"],
        "tempo_total_min": round(dur / 60, 1),
    }
    with open(os.path.join(paths["consolidado"], "relatorio_final.json"), "w",
              encoding="utf-8") as f:
        json.dump(relatorio, f, ensure_ascii=False, indent=2)
    logging.info("RELATÓRIO FINAL: %s", relatorio)

if __name__ == "__main__":
    main()
